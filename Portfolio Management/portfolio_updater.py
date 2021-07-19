# Automatic Portfolio Updater
# Connects to Coinmarketcap.com and retrieves the USD and BTC rates
# for a given set of cryptos
# Then, stores the data in the relevant cells in 'Portfolio Tracker.xlsx'

import time
import json
import pprint
import openpyxl
import requests
from requests import Request, Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell import Cell
from datetime import datetime

import wallet_balance


# Accesses coinmarketcap's API to get latest USD rates
# for each portfolio crypto
def get_usd_rates(portfolio_cryptos, api_key):
    usd_rates = {}

    for sym in portfolio_cryptos:
        url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest'
        parameters = {
            'symbol': sym
        }
        headers = {
            'Accepts': 'application/json',
            'X-CMC_PRO_API_KEY': api_key,
        }

        session = Session()
        session.headers.update(headers)

        try:
            if sym == 'USD':
                usd_rates[sym] = 1
                continue
            response = session.get(url, params=parameters)
            data = json.loads(response.text)
            usd_rates[sym] = data['data'][sym]['quote']['USD']['price']
        except (ConnectionError, Timeout, TooManyRedirects) as e:
            print(e)

    return usd_rates


# converts the usd rates into btc rates
# for each portfolio crypto
def get_btc_rates(usd_rates):
    btc_rates = {}
    btc_price = usd_rates['BTC']

    # iterate through each usd rate and
    # convert divide by btc_price
    for k, v in usd_rates.items():
        btc_rates[k] = v/btc_price

    return btc_rates


# Reads in list of symbols between startRow and endRow
def read_symbols(startRow, endRow, portfolio_tracker_file_name):
    portfolio_cryptos = []

    xfile = openpyxl.load_workbook(portfolio_tracker_file_name)
    sheet = xfile.get_sheet_by_name('Coin Holdings')

    # retrieve symbols
    for row in range(startRow, endRow):
        sym = sheet.cell(column=2, row=row).value
        if sym is not None:
            portfolio_cryptos.append(sym)

    xfile.save(portfolio_tracker_file_name)

    return portfolio_cryptos


# writes rates into appropriate cells in excel file
def write_rates(usd_rates, btc_rates, portfolio_cryptos, portfolio_tracker_file_name):
    xfile = openpyxl.load_workbook(portfolio_tracker_file_name)
    sheet = xfile.get_sheet_by_name('Coin Holdings')

    # insert usd rates
    for row in range(4, (4+len(portfolio_cryptos))):
        sym = sheet.cell(column=2, row=row).value
        _ = sheet.cell(column=4, row=row, value=usd_rates[sym])

    # insert btc rates
    for row in range(4, (4+len(portfolio_cryptos))):
        sym = sheet.cell(column=2, row=row).value
        _ = sheet.cell(column=5, row=row, value=btc_rates[sym])

    xfile.save(portfolio_tracker_file_name)


# Calcs the USD and BTC vals and sums
# (Does the excel calculations as openpyxl cannot interpret formulae)
def get_vals(usd_rates, btc_rates, sheet, portfolio_cryptos):
    usd_vals = []
    btc_vals = []
    amounts = {}

    # read amounts
    for row in range(4, (4+len(portfolio_cryptos))):
        sym = sheet.cell(column=2, row=row).value
        amount = sheet.cell(column=3, row=row).value
        amounts[sym] = amount

    # loop over cryptos and calc the $ and BTC valuations
    for i in amounts:
        usd_val = amounts[i]*usd_rates[i]
        btc_val = amounts[i]*btc_rates[i]
        usd_vals.append(usd_val)
        btc_vals.append(btc_val)

    # sum the $ and BTC valuations
    usd_sum = sum(usd_vals)
    btc_sum = sum(btc_vals)
    return usd_sum, btc_sum


# writes the latest dollar and bitcoin valuations to
# 'History Coin Holdings' and 'Daily Tracker'
def write_vals(usd_rates, btc_rates, portfolio_cryptos, portfolio_tracker_file_name):
    xfile = openpyxl.load_workbook(portfolio_tracker_file_name)

    # Open sheets
    sheet_coin_holdings = xfile.get_sheet_by_name('Coin Holdings')
    sheet_history_coin_holdings = xfile.get_sheet_by_name('History Coin Holdings')
    sheet_daily_tracker = xfile.get_sheet_by_name('Daily Tracker')

    # Now time
    now = datetime.now()
    now_formatted = now.strftime("%d/%m/%Y")

    # Calc USD and BTC vals from 'Coin Holdings'
    usd_val, btc_val = get_vals(usd_rates, btc_rates, sheet_coin_holdings, portfolio_cryptos)
    print("usd val = ", usd_val)
    print("btc val = ", btc_val)

    data = {now_formatted, btc_val, usd_val}

    # Input into 'History Coin Holdings'
    for row in range(5, 10000):
        cell_value = sheet_history_coin_holdings.cell(column=2, row=row).value
        if cell_value is None:
            _ = sheet_history_coin_holdings.cell(column=2, row=row, value=now_formatted)  # Date
            _ = sheet_history_coin_holdings.cell(column=3, row=row, value=btc_val)        # BTC Val
            _ = sheet_history_coin_holdings.cell(column=4, row=row, value=usd_val)        # USD Val
            break

    for table in sheet_history_coin_holdings.tables:
        if table == "HistoryCoinHoldings":
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            del sheet_history_coin_holdings.tables[table]
            tab = Table(displayName=table, ref=f"B4:D{row}", tableStyleInfo=style)
            sheet_history_coin_holdings.add_table(tab)

    # Input into 'Daily Tracker'
    for row in range(3, 10000):
        cell_value = sheet_daily_tracker.cell(column=2, row=row).value
        if cell_value is None:
            _ = sheet_daily_tracker.cell(column=2, row=row, value=now_formatted)          # Date
            _ = sheet_daily_tracker.cell(column=3, row=row, value=btc_val)                # BTC Val
            _ = sheet_daily_tracker.cell(column=4, row=row, value=usd_val)                # USD Val
            break

    for table in sheet_daily_tracker.tables:
        if table == "DailyTracker":
            style = TableStyleInfo(name="TableStyleMedium7", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            del sheet_daily_tracker.tables[table]
            tab = Table(displayName=table, ref=f"B2:F{row}", tableStyleInfo=style)
            sheet_daily_tracker.add_table(tab)

    xfile.save(portfolio_tracker_file_name)


# writes the balances of the cryptos into portfolio updater
def write_balances(balances, portfolio_tracker_file_name):
    xfile = openpyxl.load_workbook(portfolio_tracker_file_name)

    # Open sheets
    sheet_coin_holdings = xfile.get_sheet_by_name('Coin Holdings')
    sheet_history_coin_holdings = xfile.get_sheet_by_name('History Coin Holdings')

    # Input into 'Coin Holdings'
    for row in range(4, (4+len(balances))):
        crypto = sheet_coin_holdings.cell(column=2, row=row).value
        _ = sheet_coin_holdings.cell(column=3, row=row, value=balances[crypto])

    # Input into 'History Coin Holdings'
    for row in range(5, 10000):
        cell_value = sheet_history_coin_holdings.cell(column=2, row=row).value
        if cell_value is None:  # empty row --> insert balances here
            col = 6  # first BTC column
            for crypto in balances:
                _ = sheet_history_coin_holdings.cell(column=col, row=row, value=balances[crypto])
                col += 1
            break

    xfile.save(portfolio_tracker_file_name)


def main():
    """ You will need to change these variables as necesarry """
    # Your Coinmarketcap API key - need to make an account with them first
    api_key = ''
    # Your local file name for your portfolio tracker excel sheet
    portfolio_tracker_file_name = 'Portfolio Tracker.xlsx'

    # list of excel portfolio tickers -
    print("Retrieving crypto symbols from portfolio... \r")
    portfolio_cryptos = read_symbols(4, 100, portfolio_tracker_file_name)
    print("Successfully retrieved symbols!")

    # retrieve wallet balances from exchanges
    balances = wallet_balance.retrieve_balances(portfolio_cryptos)

    # write wallet balances to 'Coin Holdings' and 'History Coin Holdings'
    print(f"\nWriting balances to {portfolio_tracker_file_name}...")
    write_balances(balances, portfolio_tracker_file_name)
    print(f"Finished writing balances to {portfolio_tracker_file_name}!")

    # get current use rates
    print("\nAccessing coinmarketcap's api for latest USD rates... \r")
    usd_rates = get_usd_rates(portfolio_cryptos, api_key)
    print("Successfully retrieved latest USD rates!")
    print(usd_rates)

    # convert usd rates into btc rates
    print("\nAccessing coinmarketcap's api for latest BTC rates... \r")
    btc_rates = get_btc_rates(usd_rates)
    print("Successfully retrieved latest BTC rates!")
    print(btc_rates)

    # write latest rates to 'Coin Holdings' Sheet
    print(f"\nWriting rates to {portfolio_tracker_file_name}...")
    write_rates(usd_rates, btc_rates, portfolio_cryptos, portfolio_tracker_file_name)
    print(f"Finished writing rates to {portfolio_tracker_file_name}!")

    # write latest USD and BTC valuations to 'History Coin Holdings'
    # and 'Daily Tracker'
    print(f"\nWriting $ and BTC valuations to {portfolio_tracker_file_name}...")
    write_vals(usd_rates, btc_rates, portfolio_cryptos, portfolio_tracker_file_name)
    print(f"Finished writing $ and BTC valuations to {portfolio_tracker_file_name}!")

    now = datetime.now()
    now_formatted = now.strftime("%d/%m/%Y %H:%M:%S")
    print(f"\nAutomatic Update completed at {now_formatted}")


if __name__ == '__main__':
    main()
